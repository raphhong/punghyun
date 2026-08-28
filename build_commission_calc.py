# -*- coding: utf-8 -*-
"""풍현 영업 수수료 계산기 (엑셀)
- 취급액(자산 매입가)을 넣으면 수수료 총액 / 선지급 / 만기지급 자동 계산
- 요율·지급비율은 파란 글씨 셀에서 직접 조정 가능
- 구간별 시뮬 + 월수익 시나리오 시트 포함
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.properties import CalcProperties

OUT = r"C:\Users\sol hong\Downloads\풍현_영업수수료_계산기_260810.xlsx"

KF = "맑은 고딕"
NAVY = "0A1830"
GREEN = "12A06D"
LIGHT = "F4F7F6"
YELLOW = "FFF7CC"
BLUEIN = "0000FF"   # 입력/조정 셀 글씨
GREY = "5A6472"

WON = '#,##0"원"'
PCT = '0.0%'

thin = Side(style="thin", color="D5DBDA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(sz=11, b=False, color="1A2230"):
    return Font(name=KF, size=sz, bold=b, color=color)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def set_col(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

wb = openpyxl.Workbook()

# ============ 시트1: 수수료 계산기 ============
ws = wb.active
ws.title = "수수료 계산기"
ws.sheet_view.showGridLines = False
set_col(ws, {"A": 3, "B": 30, "C": 22, "D": 3, "E": 40})

# 제목
ws.merge_cells("B2:C2")
c = ws["B2"]; c.value = "풍현 영업 수수료 계산기"
c.font = F(18, True, NAVY); c.alignment = Alignment(vertical="center")
ws.row_dimensions[2].height = 30
ws.merge_cells("B3:C3")
ws["B3"].value = "취급액만 넣으면 자동 계산됩니다. 노란 칸(취급액)과 파란 글씨(요율·지급비율)만 바꾸세요."
ws["B3"].font = F(10, False, GREY)

# ① 취급액 입력
r = 5
ws.merge_cells(f"B{r}:C{r}")
ws[f"B{r}"].value = "① 취급액 입력"
ws[f"B{r}"].font = F(12, True, "FFFFFF")
ws[f"B{r}"].fill = fill(NAVY)
ws[f"C{r}"].fill = fill(NAVY)
ws[f"B{r}"].alignment = Alignment(vertical="center")
ws.row_dimensions[r].height = 22

r = 6
ws[f"B{r}"].value = "취급액 (자산 매입가)"
ws[f"B{r}"].font = F(12, True)
ws[f"B{r}"].border = border
ws[f"C{r}"].value = 1_500_000_000
ws[f"C{r}"].font = F(14, True, BLUEIN)
ws[f"C{r}"].fill = fill(YELLOW)
ws[f"C{r}"].number_format = WON
ws[f"C{r}"].border = border
ws[f"C{r}"].alignment = Alignment(horizontal="right")
ws.row_dimensions[r].height = 26
ws[f"E{r}"].value = "◀ 여기에 이번 건의 매입가를 입력하세요 (예시: 15억)"
ws[f"E{r}"].font = F(10, False, GREEN)

# ② 요율·지급 설정
r = 8
ws.merge_cells(f"B{r}:C{r}")
ws[f"B{r}"].value = "② 요율·지급 설정  (조정 가능)"
ws[f"B{r}"].font = F(12, True, "FFFFFF")
ws[f"B{r}"].fill = fill(GREEN)
ws[f"C{r}"].fill = fill(GREEN)
ws[f"B{r}"].alignment = Alignment(vertical="center")
ws.row_dimensions[r].height = 22

settings = [
    ("요율 구간 기준", 3_000_000_000, WON, "이 금액 이하/초과로 요율이 갈립니다"),
    ("요율 (기준 이하)", 0.020, PCT, "취급액이 기준 이하일 때 적용 (예: 2.0%)"),
    ("요율 (기준 초과)", 0.025, PCT, "취급액이 기준 초과일 때 적용 (예: 2.5%)"),
    ("성사 시 선지급 비율", 0.30, PCT, "총 수수료 중 성사 즉시 지급 비율 (나머지는 만기)"),
]
row0 = 9
for i, (label, val, fmt, note) in enumerate(settings):
    rr = row0 + i
    ws[f"B{rr}"].value = label
    ws[f"B{rr}"].font = F(11, True)
    ws[f"B{rr}"].border = border
    cc = ws[f"C{rr}"]
    cc.value = val
    cc.font = F(12, True, BLUEIN)
    cc.number_format = fmt
    cc.border = border
    cc.alignment = Alignment(horizontal="right")
    ws[f"E{rr}"].value = note
    ws[f"E{rr}"].font = F(9, False, GREY)

# ③ 계산 결과
r = 14
ws.merge_cells(f"B{r}:C{r}")
ws[f"B{r}"].value = "③ 계산 결과"
ws[f"B{r}"].font = F(12, True, "FFFFFF")
ws[f"B{r}"].fill = fill(NAVY)
ws[f"C{r}"].fill = fill(NAVY)
ws[f"B{r}"].alignment = Alignment(vertical="center")
ws.row_dimensions[r].height = 22

# 셀 참조: 취급액 C6, 기준 C9, 요율이하 C10, 요율초과 C11, 선지급률 C12
results = [
    ("적용 요율", "=IF(C6<=C9,C10,C11)", PCT, False, "1A2230"),
    ("수수료 총액", "=C6*IF(C6<=C9,C10,C11)", WON, True, GREEN),
    ("성사 시 선지급", "=C6*IF(C6<=C9,C10,C11)*C12", WON, True, NAVY),
    ("만기 정산 시 지급", "=C6*IF(C6<=C9,C10,C11)*(1-C12)", WON, True, NAVY),
]
row0 = 15
for i, (label, formula, fmt, big, color) in enumerate(results):
    rr = row0 + i
    ws[f"B{rr}"].value = label
    ws[f"B{rr}"].font = F(12 if big else 11, True)
    ws[f"B{rr}"].border = border
    ws[f"B{rr}"].fill = fill(LIGHT)
    cc = ws[f"C{rr}"]
    cc.value = formula
    cc.font = F(15 if big else 12, True, color)
    cc.number_format = fmt
    cc.border = border
    cc.fill = fill(LIGHT)
    cc.alignment = Alignment(horizontal="right")
    ws.row_dimensions[rr].height = 24 if big else 20

# 안내 노트
r = 20
ws.merge_cells(f"B{r}:E{r+1}")
ws[f"B{r}"].value = ("※ 참고용 계산입니다. 요율·지급비율은 예시이며 회사와의 협의·계약으로 확정됩니다. "
                     "수익을 보장하지 않으며 관련 법령 준수를 위해 조건은 변경될 수 있습니다.  대외비 · 무단 배포 금지")
ws[f"B{r}"].font = F(9, False, GREY)
ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)

# ============ 시트2: 구간별 시뮬 ============
ws2 = wb.create_sheet("구간별 시뮬")
ws2.sheet_view.showGridLines = False
set_col(ws2, {"A": 3, "B": 18, "C": 12, "D": 18, "E": 18, "F": 18})

ws2.merge_cells("B2:F2")
ws2["B2"].value = "취급액 구간별 수수료 시뮬레이션"
ws2["B2"].font = F(16, True, NAVY)
ws2.row_dimensions[2].height = 28
ws2["B3"].value = "요율·지급비율은 '수수료 계산기' 시트의 설정값을 그대로 따릅니다. 그 값을 바꾸면 여기도 자동 반영됩니다."
ws2.merge_cells("B3:F3")
ws2["B3"].font = F(9, False, GREY)

hdrs = ["취급액", "적용 요율", "수수료 총액", "성사 시 선지급", "만기 정산 지급"]
hr = 5
for j, h in enumerate(hdrs):
    cc = ws2.cell(row=hr, column=2 + j, value=h)
    cc.font = F(11, True, "FFFFFF")
    cc.fill = fill(NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = border
ws2.row_dimensions[hr].height = 22

amounts = [500_000_000, 1_000_000_000, 1_500_000_000, 2_000_000_000,
           3_000_000_000, 4_000_000_000, 5_000_000_000, 10_000_000_000]
S = "'수수료 계산기'"
for i, amt in enumerate(amounts):
    rr = hr + 1 + i
    a = f"B{rr}"
    ws2[a].value = amt
    ws2[a].number_format = WON
    ws2[a].font = F(11, True, BLUEIN)
    # 적용요율
    ws2[f"C{rr}"].value = f"=IF(B{rr}<={S}!$C$9,{S}!$C$10,{S}!$C$11)"
    ws2[f"C{rr}"].number_format = PCT
    # 총액
    ws2[f"D{rr}"].value = f"=B{rr}*C{rr}"
    ws2[f"D{rr}"].number_format = WON
    ws2[f"D{rr}"].font = F(11, True, GREEN)
    # 선지급
    ws2[f"E{rr}"].value = f"=D{rr}*{S}!$C$12"
    ws2[f"E{rr}"].number_format = WON
    # 만기
    ws2[f"F{rr}"].value = f"=D{rr}*(1-{S}!$C$12)"
    ws2[f"F{rr}"].number_format = WON
    for col in "BCDEF":
        ws2[f"{col}{rr}"].border = border
        ws2[f"{col}{rr}"].alignment = Alignment(horizontal="right")
        if i % 2 == 1:
            ws2[f"{col}{rr}"].fill = fill(LIGHT)

# ============ 시트3: 월수익 시나리오 ============
ws3 = wb.create_sheet("월수익 시나리오")
ws3.sheet_view.showGridLines = False
set_col(ws3, {"A": 3, "B": 24, "C": 20, "D": 3, "E": 16, "F": 18, "G": 18})

ws3.merge_cells("B2:G2")
ws3["B2"].value = "월 수익 시나리오"
ws3["B2"].font = F(16, True, NAVY)
ws3.row_dimensions[2].height = 28
ws3.merge_cells("B3:G3")
ws3["B3"].value = "월 몇 건을 성사시키면 얼마를 버는지 확인하세요. '건당 평균 취급액'만 바꾸면 표가 자동 갱신됩니다."
ws3["B3"].font = F(9, False, GREY)

# 입력: 건당 평균 취급액
ws3["B5"].value = "건당 평균 취급액"
ws3["B5"].font = F(12, True)
ws3["B5"].border = border
ws3["C5"].value = 1_500_000_000
ws3["C5"].font = F(12, True, BLUEIN)
ws3["C5"].fill = fill(YELLOW)
ws3["C5"].number_format = WON
ws3["C5"].border = border
ws3["C5"].alignment = Alignment(horizontal="right")

# 표 헤더
hdr3 = ["월 성사 건수", "월 취급액 합계", "월 수수료 총액", "이 중 선지급(월)"]
hr = 7
for j, h in enumerate(hdr3):
    cc = ws3.cell(row=hr, column=2 + j, value=h)
    cc.font = F(11, True, "FFFFFF")
    cc.fill = fill(GREEN)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = border
ws3.row_dimensions[hr].height = 22

counts = [1, 2, 3, 4, 5, 8, 10]
for i, n in enumerate(counts):
    rr = hr + 1 + i
    ws3[f"B{rr}"].value = n
    ws3[f"B{rr}"].number_format = '0"건"'
    ws3[f"B{rr}"].font = F(11, True, BLUEIN)
    # 월 취급액 합계 = 건수 * 건당평균
    ws3[f"C{rr}"].value = f"=B{rr}*$C$5"
    ws3[f"C{rr}"].number_format = WON
    # 월 수수료 총액 = 취급액합계 * 해당 요율 (개별 건 기준 요율 적용)
    ws3[f"D{rr}"].value = f"=C{rr}*IF($C$5<={S}!$C$9,{S}!$C$10,{S}!$C$11)"
    ws3[f"D{rr}"].number_format = WON
    ws3[f"D{rr}"].font = F(12, True, GREEN)
    # 선지급(월)
    ws3[f"E{rr}"].value = f"=D{rr}*{S}!$C$12"
    ws3[f"E{rr}"].number_format = WON
    for col in "BCDE":
        ws3[f"{col}{rr}"].border = border
        ws3[f"{col}{rr}"].alignment = Alignment(horizontal="right")
        if i % 2 == 1:
            ws3[f"{col}{rr}"].fill = fill(LIGHT)

r = hr + len(counts) + 2
ws3.merge_cells(f"B{r}:G{r+1}")
ws3[f"B{r}"].value = ("※ '월 수수료 총액'은 그 달에 성사시킨 전체 기준입니다. 실제 현금 수령은 선지급분(성사 시) + "
                      "만기 정산분(각 건의 3개월 만기 시)으로 나뉘어 들어옵니다. 참고용 예시이며 협의로 확정됩니다.")
ws3[f"B{r}"].font = F(9, False, GREY)
ws3[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)

# ============ 시트4: 회사 순이익(금액) ============
ws4 = wb.create_sheet("회사 순이익(금액)")
ws4.sheet_view.showGridLines = False
set_col(ws4, {"A": 3, "B": 32, "C": 20, "D": 3, "E": 40})

ws4.merge_cells("B2:C2")
ws4["B2"].value = "취급액 → 회사가 버는 돈 (금액)"
ws4["B2"].font = F(16, True, NAVY)
ws4.row_dimensions[2].height = 28
ws4.merge_cells("B3:E3")
ws4["B3"].value = "'수수료 계산기' 시트의 취급액·영업수수료가 자동 연동됩니다. 파란 글씨(딜 가정)만 조정하세요."
ws4["B3"].font = F(9, False, GREY)

# ① 딜 경제성 가정 (기본 시나리오 · 조정 가능)
r = 5
ws4.merge_cells(f"B{r}:C{r}")
ws4[f"B{r}"].value = "① 딜 경제성 가정  (기본 시나리오 · 조정 가능)"
ws4[f"B{r}"].font = F(12, True, "FFFFFF")
ws4[f"B{r}"].fill = fill(GREEN)
ws4[f"C{r}"].fill = fill(GREEN)
ws4[f"B{r}"].alignment = Alignment(vertical="center")
ws4.row_dimensions[r].height = 22

econ = [
    ("총마진율 (취급액 대비)", 0.21, "할부80%·렌탈20% 블렌드. 자산원가 회수 후 마진(재무모델 기본)"),
    ("유동화 수수료율", 0.075, "3개월 조달비용 (채권액면 대비)"),
    ("대손율", 0.025, "비소구 구조, 회전당"),
    ("지급수수료율", 0.005, "신탁·펌뱅킹·법무"),
    ("서비스료율", 0.003, "정산·수금 수수료 (수익, +)"),
]
row0 = 6  # 총마진율=C6, 유동화=C7, 대손=C8, 지급=C9, 서비스=C10
for i, (label, val, note) in enumerate(econ):
    rr = row0 + i
    ws4[f"B{rr}"].value = label
    ws4[f"B{rr}"].font = F(11, True)
    ws4[f"B{rr}"].border = border
    cc = ws4[f"C{rr}"]
    cc.value = val
    cc.font = F(12, True, BLUEIN)
    cc.number_format = PCT
    cc.border = border
    cc.alignment = Alignment(horizontal="right")
    ws4[f"E{rr}"].value = note
    ws4[f"E{rr}"].font = F(9, False, GREY)

# ② 이번 딜: 회사가 3개월에 버는 돈 (금액)
r = 12
ws4.merge_cells(f"B{r}:C{r}")
ws4[f"B{r}"].value = "② 이번 딜 → 회사가 3개월에 버는 돈"
ws4[f"B{r}"].font = F(12, True, "FFFFFF")
ws4[f"B{r}"].fill = fill(NAVY)
ws4[f"C{r}"].fill = fill(NAVY)
ws4[f"B{r}"].alignment = Alignment(vertical="center")
ws4.row_dimensions[r].height = 22

# 취급액=C13, 총마진=C14, 유동화=C15, 대손=C16, 지급=C17, 서비스=C18, 기여마진=C19, 영업수수료=C20, 순이익=C21, %=C22
flow = [
    ("취급액 (매입가)", f"={S}!$C$6", WON, "1A2230", False),
    ("총마진 (자산원가 회수 후)", "=C13*C6", WON, GREEN, False),
    ("(-) 유동화 수수료 (조달비용)", "=-C13*C7", WON, "C0392B", False),
    ("(-) 대손", "=-C13*C8", WON, "C0392B", False),
    ("(-) 지급수수료", "=-C13*C9", WON, "C0392B", False),
    ("(+) 서비스료", "=C13*C10", WON, GREEN, False),
    ("= 기여마진 (영업수수료 전)", "=C14+C15+C16+C17+C18", WON, NAVY, False),
    ("(-) 영업수수료", f"=-{S}!$C$16", WON, "C0392B", False),
    ("= 회사 순이익 (고정비 전)", "=C19+C20", WON, GREEN, True),
    ("취급액 대비 (3개월)", "=C21/C13", PCT, GREY, False),
]
row0 = 13
for i, (label, formula, fmt, color, big) in enumerate(flow):
    rr = row0 + i
    ws4[f"B{rr}"].value = label
    ws4[f"B{rr}"].font = F(13 if big else 11, True)
    ws4[f"B{rr}"].border = border
    cc = ws4[f"C{rr}"]
    cc.value = formula
    cc.font = F(16 if big else 12, True, color)
    cc.number_format = fmt
    cc.border = border
    cc.alignment = Alignment(horizontal="right")
    if big or label.startswith("="):
        ws4[f"B{rr}"].fill = fill(LIGHT)
        cc.fill = fill(LIGHT)
    if big:
        ws4.row_dimensions[rr].height = 26

# ③ 구간별 회사 순이익
r = 24
ws4.merge_cells(f"B{r}:E{r}")
ws4[f"B{r}"].value = "③ 취급액 구간별 회사 순이익 (영업수수료 차감 후 · 고정비 전)"
ws4[f"B{r}"].font = F(12, True, "FFFFFF")
ws4[f"B{r}"].fill = fill(NAVY)
for col in "CDE":
    ws4[f"{col}{r}"].fill = fill(NAVY)
ws4[f"B{r}"].alignment = Alignment(vertical="center")
ws4.row_dimensions[r].height = 22

set_col(ws4, {"A": 3, "B": 32, "C": 20, "D": 20, "E": 20})
hdr4 = ["취급액", "기여마진", "(-) 영업수수료", "= 회사 순이익"]
hr = 25
for j, h in enumerate(hdr4):
    cc = ws4.cell(row=hr, column=2 + j, value=h)
    cc.font = F(11, True, "FFFFFF")
    cc.fill = fill(GREEN)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = border
ws4.row_dimensions[hr].height = 22

amts4 = [500_000_000, 1_000_000_000, 1_500_000_000, 3_000_000_000, 5_000_000_000]
for i, amt in enumerate(amts4):
    rr = hr + 1 + i
    ws4[f"B{rr}"].value = amt
    ws4[f"B{rr}"].number_format = WON
    ws4[f"B{rr}"].font = F(11, True, BLUEIN)
    # 기여마진 = 취급액 * (총마진-유동화-대손-지급+서비스)
    ws4[f"C{rr}"].value = f"=B{rr}*($C$6-$C$7-$C$8-$C$9+$C$10)"
    ws4[f"C{rr}"].number_format = WON
    # 영업수수료 = 취급액 * 해당 요율 (정률, 구간별)
    ws4[f"D{rr}"].value = f"=B{rr}*IF(B{rr}<={S}!$C$9,{S}!$C$10,{S}!$C$11)"
    ws4[f"D{rr}"].number_format = WON
    # 회사 순이익
    ws4[f"E{rr}"].value = f"=C{rr}-D{rr}"
    ws4[f"E{rr}"].number_format = WON
    ws4[f"E{rr}"].font = F(11, True, GREEN)
    for col in "BCDE":
        ws4[f"{col}{rr}"].border = border
        ws4[f"{col}{rr}"].alignment = Alignment(horizontal="right")
        if i % 2 == 1:
            ws4[f"{col}{rr}"].fill = fill(LIGHT)

r = hr + len(amts4) + 2
ws4.merge_cells(f"B{r}:E{r+2}")
ws4[f"B{r}"].value = (
    "※ 위 금액은 딜 1건·3개월·기본 시나리오 기준입니다. 인건비·고정운영비(월 약 1,800만원)는 딜별이 아니라 "
    "월 단위로 별도 차감되며, 동시에 여러 건을 취급하면 건당 부담은 작아집니다.\n"
    "※ 취급액 대비 %가 낮아 보이는 것은 착시입니다 — 취급액은 대부분 유동화(외부조달)로 사는 것이고 3개월마다 회전합니다. "
    "자기자본(5.71억) 대비로는 연 37~55% 수준. 참고용이며 협의로 확정됩니다.")
ws4[f"B{r}"].font = F(9, False, GREY)
ws4[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)

# Excel이 열 때 전체 재계산
wb.calculation = CalcProperties(fullCalcOnLoad=True)
wb.save(OUT)
print("saved", OUT)

# ---- 파이썬 자체 검증 (LibreOffice 없어 recalc 불가 → 로직 재현 검증) ----
def calc(amt, thr=3e9, rlo=0.02, rhi=0.025, up=0.30):
    rate = rlo if amt <= thr else rhi
    total = amt * rate
    return rate, total, total * up, total * (1 - up)
for amt in [1_500_000_000, 3_000_000_000, 5_000_000_000, 10_000_000_000]:
    rate, tot, u, m = calc(amt)
    print(f"취급액 {amt:>14,} → 요율 {rate:.1%}, 총액 {tot:>12,.0f}, 선지급 {u:>11,.0f}, 만기 {m:>12,.0f}")

print("\n[회사 순이익(금액) 검증 · 기본 시나리오]")
GM, SEC, LOSS, PAY, SVC = 0.21, 0.075, 0.025, 0.005, 0.003
contrib_rate = GM - SEC - LOSS - PAY + SVC
for amt in [500_000_000, 1_000_000_000, 1_500_000_000, 3_000_000_000, 5_000_000_000]:
    rate, tot, u, m = calc(amt)
    contrib = amt * contrib_rate
    net = contrib - tot
    print(f"취급액 {amt:>14,} → 기여마진 {contrib:>12,.0f}, 영업수수료 {tot:>11,.0f}, 회사순이익 {net:>12,.0f}")

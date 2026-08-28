import openpyxl, sys
path = r'C:\Users\sol hong\Downloads\punghyun_invest_model_SRB_2.xlsx'
f = openpyxl.load_workbook(path)
v = openpyxl.load_workbook(path, data_only=True)
sheet = sys.argv[1] if len(sys.argv) > 1 else '가정·레버'
ws = f[sheet]; wv = v[sheet]
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c); val = wv.cell(r, c).value
        if cell.value is not None:
            print("%s: F=%r | V=%r" % (cell.coordinate, cell.value, val))

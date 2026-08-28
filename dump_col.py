import openpyxl, sys
path = r'C:\Users\sol hong\Downloads\punghyun_invest_model_SRB_2.xlsx'
f = openpyxl.load_workbook(path)
v = openpyxl.load_workbook(path, data_only=True)
sheet = sys.argv[1]
col = sys.argv[2] if len(sys.argv) > 2 else 'C'
ws = f[sheet]; wv = v[sheet]
# print label col B and the given col for every row
for r in range(1, ws.max_row + 1):
    b = ws.cell(r, 2).value
    cell = ws[col + str(r)]
    val = wv[col + str(r)].value
    if b is not None or cell.value is not None:
        print("r%s | B=%r | %s: F=%r | V=%r" % (r, b, col, cell.value, val))

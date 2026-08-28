# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

p = r'C:\Users\sol hong\Downloads\punghyun_invest_model_SRB_2.xlsx'
wb = openpyxl.load_workbook(p)

PCT = '0.0%'
MULT = '0.00"배"'
bold = Font(name='Arial', bold=True)
blk = Font(name='Arial')
grn = Font(name='Arial', color='008000')  # 시트링크
hdrfill = PatternFill('solid', fgColor='EAF2EE')

# ============ 월별 P&L: 투입자본 기준 수익률 행 추가 ============
ws = wb['월별 P&L']
# 기존: r24 영업이익률(매출대비), r25 [참고:자산], r26 신규취급, r27 운용잔고
ws['B28'] = '취급액 대비 영업이익률(사이클=3개월)'
ws['B29'] = '연환산 자본수익률(단순, ×회전)'
ws['B24'] = '영업이익률(총액매출 대비·참고)'   # 라벨 명확화
for col in range(3, 39):  # C..AL
    L = get_column_letter(col)
    ws['%s28' % L] = '=IFERROR(%s21/%s26,0)' % (L, L)   # 영업이익/신규취급
    ws['%s29' % L] = '=%s28*(12/가정·레버!$E$13)' % L    # ×(12/회전개월)
    ws['%s28' % L].number_format = PCT
    ws['%s29' % L].number_format = PCT
    ws['%s28' % L].font = blk
    ws['%s29' % L].font = blk
ws['B28'].font = bold; ws['B29'].font = bold

# ============ 연간 요약: 투자자 수익 지표 블록 추가 ============
wy = wb['연간 요약']
# 기존 마지막 데이터 r19 영업이익률. r21부터 신규 블록
wy['B21'] = '◆ 투입자본 기준 수익 지표 (총액회계 착시 보정)'
wy['B21'].font = bold
wy['B22'] = '취급액(투입자본) 합계'
wy['B23'] = '취급액 대비 영업이익률(사이클)'
wy['B24'] = '연환산 자본수익률(단순, ×회전)'
wy['B25'] = '취급액 대비 순이익률(사이클)'
for r in range(22,26):
    wy['B%d'%r].font = blk
# C9/D9/E9 = 자산 매입원가 합계 = 취급액 합계 (매입원가=취급액)
for col,c in [('C','C'),('D','D'),('E','E')]:
    wy['%s22'%col] = '=%s9'%c                        # 취급액 = 자산매입원가 합
    wy['%s23'%col] = '=IFERROR(%s16/%s22,0)'%(col,col)  # 영업이익/취급
    wy['%s24'%col] = '=%s23*(12/가정·레버!$E$13)'%col    # 연환산
    wy['%s25'%col] = '=IFERROR(%s18/%s22,0)'%(col,col)  # 순이익/취급
    wy['%s22'%col].number_format = '#,##0'
    wy['%s23'%col].number_format = PCT
    wy['%s24'%col].number_format = PCT
    wy['%s25'%col].number_format = PCT

wy['B27'] = '※ 총액회계는 자산 매입원가 회수분이 매출에 포함되어 매출 대비 이익률이 실제보다 낮게 보임. 투입자본(취급액) 대비 수익률이 실질 성과 지표.'
wy['B27'].font = Font(name='Arial', size=9, italic=True)

wb.calculation.fullCalcOnLoad = True
wb.save(p)
print('metrics added. P&L rows 24/28/29, 연간요약 21-27')
